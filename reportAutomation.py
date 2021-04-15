# program to automate daily report

from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime
from datetime import time

import os
import re

w1=load_workbook(filename=r"\\NAO.global.gmacfs.com\AllyUsers\SAT\Users\DZYDNQ\My Documents\Downloads\incident.xlsx")
# w3=load_workbook(filename=r"\\NAO.global.gmacfs.com\AllyUsers\SAT\Users\DZYDNQ\My Documents\Downloads\Agent_Info.xlsx")
# w2=load_workbook(filename=r"\\NAO.global.gmacfs.com\AllyUsers\SAT\Users\DZYDNQ\My Documents\Downloads\target.xlsx")
w2=Workbook()
s1=w1.active
s2=w2.active
# s3=w3.active
s2.title="target"
row_count = s1.max_row+1
current_month_text = datetime.now().strftime('%B')
SiteLocation=['Weaverville',
'Waterloo',
'Chandler',
'Tulsa',
'Joplin',
'Sarasota',
'Shoreview',
'Jacksonville',
'Charlotte',
'Lewisville'
]
Vendor1=['Weaverville',
'Waterloo',
'Chandler',
'Tulsa',
'Joplin',
'Sarasota']

Vendor2=['Shoreview',
'Jacksonville',
'Charlotte',
'Lewisville']
def uppercase(i):
    s = s1.cell(row=i, column=2).value
    start = s.find("Description Of Problem:") + len("Description of Problem:")
    if s.find("Description Of Problem:")<1:
        return None
    else:
        end = s.find("Agent Action Leading to Issue")
        substring = s[start:end]
        return substring

def majorelUserid(i):
    s = s1.cell(row=i, column=2).value
    start = s.find("Associate PULSE/Citrix ID:") + len("Associate PULSE/Citrix ID:")
    end = s.find("Associate Name:")
    substring = s[start:end]
    if s.find("Associate PULSE/Citrix ID:")<1:
        return None
    else:
        return substring

def allyuserid(i):
    s = s1.cell(row=i, column=2).value
    start = s.find("Associate Pulse ID:") + len("Associate Pulse ID:")
    end = s.find("Premier Associate:")
    substring = s[start:end]
    if s.find("Premier Associate:")<1:
        return None
    else:
        return substring
        
def allyName():
    s = s1.cell(row=i, column=2).value
    start = s.find("Associate Name:") + len("Associate Name:")
    end = s.find("Computer Name or Workstation IP:")
    substring = s[start:end]
    if s.find("Computer Name or Workstation IP:")<1:
        return None
    else:
        return substring
# def App_issueCategory(z):        
#         if 'ATTEMPTING TO SAVE THE SESSION' or 'INTERACTION' in z:
#             x='Salesforce'
#             y='Save and End Session'
#             return x, y
#         elif 'WRAP' in z:
#             x='Salesforce'
#             y='Call Wrap'
#             return x, y
#         elif 'CD FUNDS' or 'CD MANAGEMENT' or 'FUNDS' or 'WORKFLOW' in z:
#             x='WebCSR'
#             y='Native App'
#             return x, y
#         elif 'MESSAGINGHOST' or '45' or 'FAILED TO CONNECT' or 'embedded page' in z:
#             x='Salesforce'
#             y='Login Error'
#             return x, y
#         elif 'UNREPSONSIVE' or 'LATENCY' in z :
#             x='Salesforce'
#             y='Latency'
#             return x, y
#         elif 'AUTHENTICATING CUSTOMER' in z:
#             x='Salesforce'
#             y='Authentication'
#             return x, y
#         elif  'EFT' in z:
#             x='EFT'
#             y='Open Span Crash Defect'
#         elif  'CASHEDGE' in z:
#             x='CashEdge'
#             y='Open Span Crash Defect'
#             return x, y
#         elif  'BLANK IE' in z:
#             x='IE'
#             y='Blank Page'
#             return x, y
#         elif 'LN' or 'LEXIS' or 'NEXIS' or 'LEX' in z:
#             x='Lexis Nexis'
#             y='Open Span Crash Defect'
#             return x, y
#         elif 'CEDAR' in z:
#             x= 'Cedar'
#             y= 'Open Span Crash Defect'
#             return x, y
#         elif 'SYNC WEBCSR' or 'SYNC' in z:
#             x='WebCSR'
#             y='Latency'
#             return x, y
#         elif 'SEARCH FOR CUSTOMER' in z:
#             x='Salesforce'
#             y='Customer Search'
#             return x, y
#         elif 'SPINNING' or 'WHEEL' in z:
#             x='Salesforce'
#             y='Spinning Wheel'
#             return x, y
#         elif 'COMMUNICATION ERROR' or 'COMMUNICATION' in z:
#             x='Salesforce'
#             y='Communication Error'
#             return x, y
#         elif 'FREEZE' or 'CRASH' in z:
#             x='Salesforce'
#             y='Freeze/Crash'
#             return x, y
#         elif 'AVAYA' or 'MUTE' or 'HEAR' or 'SOUND' or 'HEADPHONE' or \
#             'EXTENSION' or 'TALKING' or 'AUDIO' or 'MICROPHONE' or 'SPEAKER' or \
#             'DIAL' or 'HEADSET' or 'MUSIC' or 'AUX' or 'ACW' or 'CTI' or 'GHOST CALLS' or \
#             'CALLS' or 'CALL' or 'GHOST' or 'ECHO' or 'MUFFLED' or \
#             'HEAD' or 'SOFTPHONE' or 'SOFT' or 'PHONE' or  'ROBOTIC' in z:
#             x='Avaya'
#             y='Telephony'
#             return x, y
#         elif 'VPN' or 'NETWORK' in z:
#             x='Avaya'
#             y='Local IT'
#             return x, y
#         elif 'BLANK IE' or 'BLANK PAGE' or 'INTERNET EXPLORER' or 'IE' or 'BLANK HTML IE' or 'BLANK SCREEN' or \
#             'WHITE SCREEN' in z:
#             x='IE'
#             y='Blank Page'
#             return x, y
#         else:
#             return None
        
        
for i in range (2, row_count):
  s2.cell(row=i, column=1).value=i-1
  s2.cell(row=i, column=2).value='Sev 4'
  s2.cell(row=i, column=3).value=s1.cell(row=i, column=1).value
  s2.cell(row=i, column=4).value=current_month_text
  s2.cell(row=i, column=5).value=s1.cell(row=i, column=3).value.strftime("%m/%d/%Y")
  s2.cell(row=i, column=6).value=s1.cell(row=i, column=3).value.strftime("%#I:%M:%S %p")
  s2.cell(row=i, column=7).value='Open'
  s2.cell(row=i, column=11).value=1
s2.cell(row=1, column=1).value='S.No'
s2.cell(row=1, column=2).value='Sev'
s2.cell(row=1, column=3).value='Incident'
s2.cell(row=1, column=4).value='Month'
s2.cell(row=1, column=5).value='Reported Date'
s2.cell(row=1, column=6).value='Reported Time '
s2.cell(row=1, column=7).value='Status'
s2.cell(row=1, column=8).value='Closed Date'
s2.cell(row=1, column=9).value='Site Location'
s2.cell(row=1, column=10).value='Vendor'
s2.cell(row=1, column=11).value='Agents Impacted'
s2.cell(row=1, column=12).value='Application Impacted'
s2.cell(row=1, column=13).value='Issue Category'
s2.cell(row=1, column=14).value='Problem Summary'
s2.cell(row=1, column=15).value='Agent Name'
s2.cell(row=1, column=16).value='Pulse ID / Citrix ID'
for i in range (2, row_count):
    if s1.cell(row=i, column=2).value is None:
        s1.cell(row=i, column=2).value='test'
        
#fills the site location
for i in range (2, row_count):
    for item in  SiteLocation:
        if item in s1.cell(row=i, column=2).value:
            s2.cell(row=i, column=9).value=item
        else:
            continue
            
for i in range (2, row_count):
    if s2.cell(row=i, column=9).value is None:
        s2.cell(row=i, column=9).value='test'
    else:
        continue
        
#fills the vendor type
for i in range (2, row_count):
        for item in SiteLocation:
            if item in s2.cell(row=i, column=9).value:
                if item in Vendor1:
                    s2.cell(row=i, column=10).value='3rd Party Supplier'
                elif item in Vendor2:
                    s2.cell(row=i, column=10).value='Ally Internal'
            else:
                continue
                
for i in range (2, row_count):
    if s2.cell(row=i, column=9).value is 'test':
        s2.cell(row=i, column=9).value=None
        
for i in range (2, row_count):
     s = s1.cell(row=i, column=2).value
     start = s.find("Description of Problem:") + len("Description of Problem:")
     end = s.find("Agent Action Leading to Issue")
     substring = s[start:end]
     if s.find("Description of Problem:")<1:
        s2.cell(row=i, column=14).value=uppercase(i)
        if uppercase(i) is None:
            s2.cell(row=i, column=14).value=s1.cell(row=i, column=2).value
     else:
        s2.cell(row=i, column=14).value=substring   
     
    
for i in range (2, row_count):
    go=s2.cell(row=i, column=14).value
    bo=go.lstrip()
    s2.cell(row=i, column=14).value=bo
        
for i in range (2, row_count):
    s = s1.cell(row=i, column=2).value
    start = s.find("Associate Name:") + len("Associate Name:")
    end = s.find("Citrix Server ID:")
    substring = s[start:end]
    if s.find("Citrix Server ID:")<1:
        s2.cell(row=i, column=16).value=allyName()
    else:
        s2.cell(row=i, column=16).value=substring 
        
for i in range (2, row_count):
    x=s2.cell(row=i, column=16).value
    if x is None:
        continue
    else:
        s=s2.cell(row=i, column=16).value
        s=re.sub(r'[^\w\s]','',str(s))
        s2.cell(row=i, column=16).value=s
    
    
for i in range (2, row_count):
    s = s1.cell(row=i, column=2).value
    start = s.find("Associate Pulse ID:") + len("Associate Pulse ID:")
    end = s.find("Associate Name:")
    substring = s[start:end]
    if s.find("Associate Pulse ID:")<1:
        s2.cell(row=i, column=17).value=allyuserid(i)
        if allyuserid(i) is None:
            s2.cell(row=i, column=17).value=majorelUserid(i)
    else:
        s2.cell(row=i, column=17).value=substring
    
            
for i in range (2, row_count):
    x=s2.cell(row=i, column=17).value
    if x is None:
        continue
    else:
        y=x.lstrip()
        s2.cell(row=i, column=17).value=y[0:6]    
        
for i in range (2, row_count):
    x=s2.cell(row=i, column=16).value
    if x is None:
        continue
    else:
        y=x.lstrip()
        s2.cell(row=i, column=16).value=y    
        
for i in range (2, row_count):
    x=s2.cell(row=i, column=16).value
    if x is None:
        s2.cell(row=i, column=16).value='Guest'
    elif len(x)==0:
        s2.cell(row=i, column=16).value='Guest'        
    else:
        continue
        
for i in range (2, row_count):
    x=s2.cell(row=i, column=17).value
    if x is None:
        s2.cell(row=i, column=17).value='Guest'
    elif len(x)==0:
        s2.cell(row=i, column=17).value='Guest'        
    else:
        continue        
for i in range (2, row_count):
    date_time_str = s2.cell(row=i, column=5).value
    date_time_obj = datetime.strptime(date_time_str, '%m/%d/%Y')
    s2.cell(row=i, column=5).value=date_time_obj.date()

       
for i in range (2, row_count):
    s2.cell(row=i, column=6).value=s1.cell(row=i, column=3).value.strftime("%m/%d/%Y %I:%M:%S %p")
    date_time_str = s2.cell(row=i, column=6).value
    date_time_obj1 = datetime.strptime(date_time_str, '%m/%d/%Y %I:%M:%S %p')
    s2.cell(row=i, column=6).value=date_time_obj1.time()
    

    
w5=load_workbook(filename=r"\\NAO.global.gmacfs.com\AllyUsers\SAT\Users\DZYDNQ\My Documents\Downloads\moo.xlsx")
s5=w5.active
for i in range (2, row_count):
    
    if s2.cell(row=i, column=16).value == 'Guest' and s2.cell(row=i, column=17).value !='Guest':
        for x in range (2, 1612):
            if s5.cell(row=x, column=4).value==s2.cell(row=i, column=17).value.upper():
                s2.cell(row=i, column=16).value=s5.cell(row=x, column=3).value
                s2.cell(row=i, column=9).value=s5.cell(row=x, column=1).value
                s2.cell(row=i, column=10).value=s5.cell(row=x, column=2).value
            
            
    elif s2.cell(row=i, column=16).value == 'Guest' and s2.cell(row=i, column=17).value =='Guest':
                s2.cell(row=i, column=16).value=s1.cell(row=i, column=4).value
                for x in range (2, 1612):
                    v=s2.cell(row=i, column=16).value.upper()
                    if s5.cell(row=x, column=3).value==v:
                        s2.cell(row=i, column=17).value=s5.cell(row=x, column=4).value
                        s2.cell(row=i, column=9).value=s5.cell(row=x, column=1).value
                        s2.cell(row=i, column=10).value=s5.cell(row=x, column=2).value
                    else:
                        continue
    
    
for i in range (2, row_count):
        z=s2.cell(row=i, column=14).value.upper()
        if '45' in z:
           s2.cell(row=i, column=12).value='Salesforce'
           s2.cell(row=i, column=13).value='Login Error'
        elif 'MESSAGING' in z:
           s2.cell(row=i, column=12).value='Salesforce'
           s2.cell(row=i, column=13).value='Login Error'
        
for i in range (2, row_count):
        z=s2.cell(row=i, column=14).value.upper()
        if s2.cell(row=i, column=12).value is None:
             if 'VPN' in z:
                s2.cell(row=i, column=12).value='VPN'
                s2.cell(row=i, column=13).value='Local IT'
for i in range (2, row_count):
        z=s2.cell(row=i, column=14).value.upper()
        if s2.cell(row=i, column=12).value is None:
            if 'COMMUNICATION' in z:
                s2.cell(row=i, column=12).value='Salesforce'
                s2.cell(row=i, column=13).value='Communication Error'
for i in range (2, row_count):
        z=s2.cell(row=i, column=14).value.upper()
        if s2.cell(row=i, column=12).value is None:
            if 'SESSION' in z:
                if 'SAVE' in z:
                    s2.cell(row=i, column=12).value='Salesforce'
                    s2.cell(row=i, column=13).value='Save and End Session'
for i in range (2, row_count):
        z=s2.cell(row=i, column=14).value.upper()
        if s2.cell(row=i, column=12).value is None:
            if 'AVAYA' in z:
                s2.cell(row=i, column=12).value='Avaya'
                s2.cell(row=i, column=13).value='Telephony'
            elif 'HEAR' in z:
                s2.cell(row=i, column=12).value='Avaya'
                s2.cell(row=i, column=13).value='Telephony'
            elif 'ECHO' in z:
                s2.cell(row=i, column=12).value='Avaya'
                s2.cell(row=i, column=13).value='Telephony'
            elif 'HEADSET' in z:
                s2.cell(row=i, column=12).value='Avaya'
                s2.cell(row=i, column=13).value='Telephony'
            elif 'SOFT' in z:
                s2.cell(row=i, column=12).value='Avaya'
                s2.cell(row=i, column=13).value='Telephony'
            elif 'MUTE' in z:
                s2.cell(row=i, column=12).value='Avaya'
                s2.cell(row=i, column=13).value='Telephony'
            elif 'HEADPHONE' in z:
                s2.cell(row=i, column=12).value='Avaya'
                s2.cell(row=i, column=13).value='Telephony'
            elif 'ROBOTIC' in z:
                s2.cell(row=i, column=12).value='Avaya'
                s2.cell(row=i, column=13).value='Telephony'
            elif 'GHOST' in z:
                s2.cell(row=i, column=12).value='Avaya'
                s2.cell(row=i, column=13).value='Telephony'
            elif 'MUSIC' in z:
                s2.cell(row=i, column=12).value='Avaya'
                s2.cell(row=i, column=13).value='Telephony'
            elif 'AUX' in z:
                s2.cell(row=i, column=12).value='Avaya'
                s2.cell(row=i, column=13).value='Telephony'
            elif 'ACW' in z:
                s2.cell(row=i, column=12).value='Avaya'
                s2.cell(row=i, column=13).value='Telephony'
            elif 'SILENT' in z:
                s2.cell(row=i, column=12).value='Avaya'
                s2.cell(row=i, column=13).value='Telephony'
            elif ' RING' in z:
                s2.cell(row=i, column=12).value='Avaya'
                s2.cell(row=i, column=13).value='Telephony'
for i in range (2, row_count):
        z=s2.cell(row=i, column=14).value.upper()
        if s2.cell(row=i, column=12).value is None:
            if 'WRAP' in z:
                s2.cell(row=i, column=12).value='Salesforce'
                s2.cell(row=i, column=13).value='Call Wrap'
for i in range (2, row_count):
        z=s2.cell(row=i, column=14).value.upper()
        if s2.cell(row=i, column=12).value is None:
            if 'LEX' in z:
                s2.cell(row=i, column=12).value='Lexis Nexis'
                s2.cell(row=i, column=13).value='Open Span Crash Defect'
for i in range (2, row_count):
        z=s2.cell(row=i, column=14).value.upper()
        if s2.cell(row=i, column=12).value is None:
            if 'OPENSPAN' in z:
                s2.cell(row=i, column=12).value='Salesforce'
                s2.cell(row=i, column=13).value='Open Span Crash Defect'
            elif 'OPEN SPAN' in z:
                s2.cell(row=i, column=12).value='Salesforce'
                s2.cell(row=i, column=13).value='Open Span Crash Defect'
for i in range (2, row_count):
        z=s2.cell(row=i, column=14).value.upper()
        if s2.cell(row=i, column=12).value is None:
            if 'SPINNING' in z:
                s2.cell(row=i, column=12).value='Salesforce'
                s2.cell(row=i, column=13).value='Spinning Wheel'
for i in range (2, row_count):
        z=s2.cell(row=i, column=14).value.upper()
        if s2.cell(row=i, column=12).value is None:
            if 'BLANK IE' in z:
                s2.cell(row=i, column=12).value='IE'
                s2.cell(row=i, column=13).value='Blank Page'
for i in range (2, row_count):
        z=s2.cell(row=i, column=14).value.upper()
        if s2.cell(row=i, column=12).value is None:
            if 'WEBCSR' in z:
                if 'TIMEOUT' in z:
                    s2.cell(row=i, column=12).value='WebCSR'
                    s2.cell(row=i, column=13).value='Time Out'
                elif 'FROZE' in z:
                    s2.cell(row=i, column=12).value='WebCSR'
                    s2.cell(row=i, column=13).value='Freeze/Crash'
                elif 'LOGOUT' in z:
                    s2.cell(row=i, column=12).value='WebCSR'
                    s2.cell(row=i, column=13).value='Native App'
                elif 'TIME' in z:
                    s2.cell(row=i, column=12).value='WebCSR'
                    s2.cell(row=i, column=13).value='Time Out'
for i in range (2, row_count):
        z=s2.cell(row=i, column=14).value.upper()
        if s2.cell(row=i, column=12).value is None:
            if 'SALESFORCE' in z:
                if 'TIME' in z:
                    s2.cell(row=i, column=12).value='Salesforce'
                    s2.cell(row=i, column=13).value='Time Out'
for i in range (2, row_count):
        z=s2.cell(row=i, column=14).value.upper()
        if s2.cell(row=i, column=12).value is None:
            if 'CCID' in z:
                if 'FROZE' in z:
                    s2.cell(row=i, column=12).value='CCID2'
                    s2.cell(row=i, column=13).value='Freeze/Crash'
                elif 'LOAD' in z:
                        s2.cell(row=i, column=12).value='CCID2'
                        s2.cell(row=i, column=13).value='Freeze/Crash'
                elif 'LATENCY' in z:
                            s2.cell(row=i, column=12).value='CCID2'
                            s2.cell(row=i, column=13).value='Freeze/Crash'
                elif 'CRASH' in z:
                           s2.cell(row=i, column=12).value='CCID2'
                           s2.cell(row=i, column=13).value='Freeze/Crash'     
                elif 'FREEZE' in z:
                             s2.cell(row=i, column=12).value='CCID2'
                             s2.cell(row=i, column=13).value='Freeze/Crash'
                elif 'FAIL' in z:
                             s2.cell(row=i, column=12).value='CCID2'
                             s2.cell(row=i, column=13).value='Freeze/Crash'
for i in range (2, row_count):
        z=s2.cell(row=i, column=14).value.upper()
        if s2.cell(row=i, column=12).value is None:
            if 'WORKFLOW' in z:
                s2.cell(row=i, column=12).value='WebCSR'
                s2.cell(row=i, column=13).value='Native App'
            elif 'CD' in z:
                s2.cell(row=i, column=12).value='WebCSR'
                s2.cell(row=i, column=13).value='Native App'
            elif 'FUNDS' in z:
                s2.cell(row=i, column=12).value='WebCSR'
                s2.cell(row=i, column=13).value='Native App'
for i in range (2, row_count):
        z=s2.cell(row=i, column=14).value.upper()
        if s2.cell(row=i, column=12).value is None:
            if 'DELUXE' in z:
                s2.cell(row=i, column=12).value='Deluxe'
                s2.cell(row=i, column=13).value='Native App'
for i in range (2, row_count):
        z=s2.cell(row=i, column=14).value.upper()
        if s2.cell(row=i, column=12).value is None:
            if 'CEDAR' in z:
                s2.cell(row=i, column=12).value='Cedar'
                s2.cell(row=i, column=13).value='Native App'
for i in range (2, row_count):
        z=s2.cell(row=i, column=14).value.upper()
        if s2.cell(row=i, column=12).value is None:
            if 'FISERV' in z:
                s2.cell(row=i, column=12).value='Fiserv'
                s2.cell(row=i, column=13).value='Native App'
for i in range (2, row_count):
        z=s2.cell(row=i, column=14).value.upper()
        if s2.cell(row=i, column=12).value is None:
            if ' EFT' in z:
                s2.cell(row=i, column=12).value='EFT'
                s2.cell(row=i, column=13).value='Native App'
for i in range (2, row_count):
        z=s2.cell(row=i, column=14).value.upper()
        if s2.cell(row=i, column=12).value is None:
            if 'ASKIVA' in z:
                s2.cell(row=i, column=12).value='Salesforce'
                s2.cell(row=i, column=13).value='Password Reset'
for i in range (2, row_count):
        z=s2.cell(row=i, column=14).value.upper()
        if s2.cell(row=i, column=12).value is None:
            if 'SF' in z:
                if 'TIMEOUT' in z:
                    s2.cell(row=i, column=12).value='Salesforce'
                    s2.cell(row=i, column=13).value='Time Out'
for i in range (2, row_count):
        z=s2.cell(row=i, column=14).value.upper()
        if s2.cell(row=i, column=12).value is None:
            if 'LATENCY' in z:
                if 'SYSTEM' in z:
                    s2.cell(row=i, column=12).value='Desktop'
                    s2.cell(row=i, column=13).value='Latency'
                elif 'SF' in z:
                    s2.cell(row=i, column=12).value='Salesforce'
                    s2.cell(row=i, column=13).value='Latency'
                elif 'SALESFORCE' in z:
                    s2.cell(row=i, column=12).value='Salesforce'
                    s2.cell(row=i, column=13).value='Latency'
                elif 'WEBCSR' in z:
                    s2.cell(row=i, column=12).value='WebCSR'
                    s2.cell(row=i, column=13).value='Latency'
                else:
                    s2.cell(row=i, column=13).value='Latency'
# for i in range (2, row_count):
#         z=s2.cell(row=i, column=14).value.upper()
#         if s2.cell(row=i, column=12).value is None:
#             if 'AVAYA' or 'MUTE' or 'HEAR' or 'SOUND' or 'HEADPHONE' or \
#             'EXTENSION' or 'TALKING' or 'AUDIO' or 'MICROPHONE' or 'SPEAKER' or \
#             'DIAL' or 'HEADSET' or 'MUSIC' or 'AUX' or 'ACW' or 'CTI' or 'GHOST CALLS' or \
#             'CALLS' or 'CALL' or 'GHOST' or 'ECHO' or 'MUFFLED' or \
#             'HEAD' or 'SOFTPHONE' or 'SOFT' or 'PHONE' or  'ROBOTIC' in z:
#                 s2.cell(row=i, column=12).value=='Avaya'
#                 s2.cell(row=i, column=13).value='Telephony'             
# for i in range (2, row_count):
#         z=s2.cell(row=i, column=14).value.upper()
#         if s2.cell(row=i, column=12).value is None:
#             if 'BLANK IE' or 'BLANK PAGE' or 'INTERNET EXPLORER' or 'IE' or 'BLANK HTML IE' or 'BLANK SCREEN' or \
#             'WHITE SCREEN' in z:
#                 s2.cell(row=i, column=12).value='IE'
#                 s2.cell(row=i, column=13).value='Blank Page'
# for i in range (2, row_count):
#         z=s2.cell(row=i, column=14).value.upper()
#         if s2.cell(row=i, column=12).value is None:
#             if 'ATTEMPTING TO SAVE THE SESSION' or 'INTERACTION' in z:
#                 s2.cell(row=i, column=12).value='Salesforce'
#                 s2.cell(row=i, column=13).value='Save and End Session'
# for i in range (2, row_count):
#         z=s2.cell(row=i, column=14).value.upper()
#         if s2.cell(row=i, column=12).value is None:
#             if 'CD FUNDS' or 'CD MANAGEMENT' or 'FUNDS' or 'WORKFLOW' in z:
#                 s2.cell(row=i, column=12).value='WebCSR'
#                 s2.cell(row=i, column=13).value='Native App'
# for i in range (2, row_count):
#         z=s2.cell(row=i, column=14).value.upper()
#         if s2.cell(row=i, column=12).value is None:
#             if 'WRAP' in z:
#                 s2.cell(row=i, column=12).value='Salesforce'
#                 s2.cell(row=i, column=13).value='Call Wrap'
# for i in range (2, row_count):
#         z=s2.cell(row=i, column=14).value.upper()
#         if s2.cell(row=i, column=12).value is None:
#             if 'LN' or 'LEXIS' or 'NEXIS' or 'LEX' in z:
#                 s2.cell(row=i, column=12).value='Lexis Nexis'
#                 s2.cell(row=i, column=13).value='Open Span Crash Defect'
# for i in range (2, row_count):
#         z=s2.cell(row=i, column=14).value.upper()
#         if s2.cell(row=i, column=12).value is None:
#             if 'SPINNING' or 'WHEEL' in z:
#                 s2.cell(row=i, column=12).value='Salesforce'
#                 s2.cell(row=i, column=13).value='Spinning Wheel'
# for i in range (2, row_count):
#         z=s2.cell(row=i, column=14).value.upper()
#         if s2.cell(row=i, column=12).value is None:
#             if 'WEBCSR' not in z:
#                 if 'SALESFORCE' or 'SF' in z:
#                     if 'UNREPSONSIVE' or 'LATENCY' or 'FREEZE' in z:
#                         s2.cell(row=i, column=12).value='Salesforce'
#                         s2.cell(row=i, column=13).value='Latency'                                      
# for i in range (2, row_count):
#         z=s2.cell(row=i, column=14).value.upper()
#         if s2.cell(row=i, column=12).value is None:
#             if 'SALESFORCE' not in z:
#                     if 'WEBCSR' in z:
#                         if 'SYNC WEBCSR' or 'SYNC' or 'FREEZE' or 'CRASH' or 'LATENCY' or 'LOAD' or 'UNRESPONSIVE'  in z:
#                             s2.cell(row=i, column=12).value='WebCSR'
#                             s2.cell(row=i, column=13).value='Latency'
# for i in range (2, row_count):
#         z=s2.cell(row=i, column=14).value.upper()
#         if s2.cell(row=i, column=12).value is None:
#             if 'AUTHENTICATING CUSTOMER' in z:
#                 s2.cell(row=i, column=12).value='Salesforce'
#                 s2.cell(row=i, column=13).value='Authentication'
# for i in range (2, row_count):
#         z=s2.cell(row=i, column=14).value.upper()
#         if s2.cell(row=i, column=12).value is None:
#             if 'SEARCH FOR CUSTOMER' in z:
#                 s2.cell(row=i, column=12).value='Salesforce'
#                 s2.cell(row=i, column=13).value='Customer Search'
# for i in range (2, row_count):
#         z=s2.cell(row=i, column=14).value.upper()
#         if s2.cell(row=i, column=12).value is None:
#             if 'EFT' and 'OPENSPAN' in z:
#                 s2.cell(row=i, column=12).value='EFT'
#                 s2.cell(row=i, column=13).value='Open Span Crash Defect'
# for i in range (2, row_count):
#         z=s2.cell(row=i, column=14).value.upper()
#         if s2.cell(row=i, column=12).value is None:
#             if 'EFT' and 'OPENSPAN' in z:
#                 s2.cell(row=i, column=12).value='EFT'
#                 s2.cell(row=i, column=13).value='Open Span Crash Defect'               
# for i in range (2, row_count):
#         z=s2.cell(row=i, column=14).value.upper()
#         if s2.cell(row=i, column=12).value is None:
#             if 'CEDAR' and 'OPENSPAN' in z:
#                 s2.cell(row=i, column=12).value= 'Cedar'
#                 s2.cell(row=i, column=13).value= 'Open Span Crash Defect'               
                
                
                
        
w2.save(filename=r"\\NAO.global.gmacfs.com\AllyUsers\SAT\Users\DZYDNQ\My Documents\Downloads\target.xlsx")
w2.close()
w5.close()
os.remove(r"\\NAO.global.gmacfs.com\AllyUsers\SAT\Users\DZYDNQ\My Documents\Downloads\incident.xlsx")
os.startfile(r"\\NAO.global.gmacfs.com\AllyUsers\SAT\Users\DZYDNQ\My Documents\Downloads\target.xlsx")
