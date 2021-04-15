# program to automate bulk email sending from a excel data source

from openpyxl import Workbook

from openpyxl import load_workbook

import smtplib

import re

import os

import time

w1=load_workbook(r"C:\Users\566251\Downloads\Gmail Leads.xlsx")
# w2=load_workbook(r"C:\Users\566251\Downloads\newGmailLeads.xlsx")

gmailLeads=w1.active
# newGmailLeads=w2.active

maxRow=gmailLeads.max_row+1

# mailObj=smtplib.SMTP('smtp.gmail.com', 587)
# mailObj.login('adam.neobds@gmail.com', 'Berich@3')
# mailObj.starttls()
mailObj=smtplib.SMTP_SSL('smtp.gmail.com', 465)
mailObj.login('adam.neobds@gmail.com', 'Berich@3')
content="""Subject: {}\n{}"""
txt1="""Hi {},
 
Do you leverage LinkedIn for prospecting, we can help you build a list of your future customers accompanied with their business email address & phone numbers.
 
We build your list using our own AI technology with guaranteed 95% accuracy. Regardless of how our  prospect wants the lists – to increase the probability of a deal we find lists with KEYWORDS/GROUPS/Likes/etc.
 
Would you like to discuss over a 15-minute online demo? Please let me know your convenient date and time.
 
I can demonstrate how we build your ideal contacts and accounts and give you 50 qualified leads as a thank you.

Thanks,

Adam Levine| Neo Business Development Solutions.
Business Development Manager|
P: 1415-231-3714"""
txt2="""A unique route to Increase sales for {}"""
a=0
while 1:
    for i in range(1, maxRow):
        print(i)
        if gmailLeads.cell(row=i, column=6).value=='Done':
            continue
        elif gmailLeads.cell(row=i, column=6).value is None:
            emailBody=txt1.format((gmailLeads.cell(row=i, column=1).value).split()[0])
            companyName=gmailLeads.cell(row=i, column=3).value
            if 'ï¿½ LLC' in companyName:
                subj=txt2.format(companyName.replace('ï¿½ LLC', ''))
            elif 'ï¿½ Inc' in companyName:
                subj=txt2.format(companyName.replace('ï¿½ Inc', ''))
            else:
                subj=txt2.format(companyName)
#             print(subj)
#             print(emailBody)
#             print('*****************************************************************************')
            emailTo=gmailLeads.cell(row=i, column=4).value
            x=content.format(subj, emailBody)
            mailObj.sendmail('adam.neobds@gmail.com', emailTo, str(x).encode('utf-8'))
            gmailLeads.cell(row=i, column=6).value='Done'
            a+=1
            time.sleep(60)
            if a == 10:
                break
    if a == 10:
            break
w1.save(r"C:\Users\566251\Downloads\Gmail Leads.xlsx")            
os.startfile(r"C:\Users\566251\Downloads\Gmail Leads.xlsx")
mailObj.sendmail('adam.neobds@gmail.com', 'neobdsol@gmail.com', str(x).encode('utf-8'))
mailObj.quit()
